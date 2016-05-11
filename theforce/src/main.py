# -*- coding: utf-8 -*-
import sys
import os.path
import config
import itertools
import operator

import MySQLdb
from openpyxl import load_workbook
from pymongo.mongo_client import MongoClient

import metafile
from process import process, process_ext


def load_data(meta, xlsx_file, xlsx_config):
    """
    生成器，从excel中读取数据，以region为一组返回
    """
    workbook = load_workbook(filename=xlsx_file,
                             data_only=True,
                             read_only=True)

    for sheet_name, regions in meta.iter_sheets(xlsx_config):
        sheet = workbook.get_sheet_by_name(sheet_name)
        if not sheet:
            raise Exception(u'在文件 %s 未找到Sheet: %s' % (xlsx_file, sheet_name))
        else:
            for region, item_model, extend in meta.iter_regions(regions):
                print u'准备读取 %s 区域: %s' % (sheet_name, region)
                items = []
                for row in sheet.iter_rows(region):
                    item = extend.copy() if extend else {}
                    row_no = 0

                    names = meta.get_model_attr_names(item_model)
                    values = []
                    for col in row:
                        values.append(col.value)
                        row_no = col.row

                    item['__src'] = xlsx_file
                    item['__src_sheet'] = sheet_name
                    item['__src_sheet_row'] = row_no

                    item.update(dict([(name, value)
                                      for name, value in itertools.izip(
                                          names, values)]))

                    items.append(item)
                yield item_model, items


def scan_data_dir(paths):
    """
    解析传入目录下的所有带meta.yml
    """
    paths = [_path if os.path.isabs(_path) else os.path.abspath(_path)
             for _path in paths]

    dirnames = set()

    def visit(arg, dirname, names):
        if os.path.isfile(os.path.join(dirname, 'meta.yml')):
            dirnames.add(dirname)

    for path in paths:
        os.path.walk(path, visit, {})

    return [metafile.Meta(os.path.join(dirname, 'meta.yml'))
            for dirname in dirnames]


def load(metas):
    results = []

    for meta in metas:
        print u'解析目录: ', meta._dirname
        for xlsx_file, xlsx_config in meta.iter_xlsx():
            print u'解析Excel文件:', xlsx_file
            for item_model, items in load_data(meta, xlsx_file, xlsx_config):
                objs = []
                for obj in items:
                    obj, errors, warnings = process(meta, obj, item_model)
                    obj['__errors'] = errors
                    obj['__warnings'] = warnings

                    obj, errors, warnings = process_ext(meta, obj, item_model)
                    obj['__errors'].extend(errors)
                    obj['__warnings'].extend(warnings)
                    objs.append(obj)
                results.append([item_model, objs])

    error_count = reduce(
        operator.add,
        [len(obj['__errors'])
         for obj in itertools.chain(*[objs for item_model, objs in results])])
    warning_count = reduce(
        operator.add,
        [len(obj['__warnings'])
         for obj in itertools.chain(*[objs for item_model, objs in results])])

    error_result_count = reduce(
        operator.add,
        [1
         for obj in itertools.chain(*[objs for item_model, objs in results])
         if len(obj['__errors']) > 0], 0)
    result_count = reduce(operator.add, [len(objs)
                                         for item_model, objs in results])

    def print_summary():
        print
        print "=" * 40
        print ' ' * 15, u'Summary'
        print "=" * 40
        print
        print "数据量:", result_count
        print "包含错误的数据量:", error_result_count
        print "错误:", error_count
        print "警告:", warning_count

    if error_count > 0:
        # 有错误，需先纠错
        for obj in itertools.chain(*[objs for item_model, objs in results]):
            if len(obj['__errors']) > 0:
                print '-' * 40
                print obj['__src'], obj['__src_sheet'], obj['__src_sheet_row']
                print '\n'.join(obj['__errors'])
    print_summary()

    return results


def stash(results):
    """
    暂存到mongo数据库中。
    """
    summary = {}
    mongo = MongoClient(**config.mongo)
    try:
        for item_model, objs in results:
            collection_name = item_model['name']
            db = mongo.get_database('theforce')
            collection = db.get_collection(collection_name)
            collection.insert_many(objs)
            summary[collection_name] = len(
                objs) if collection_name not in summary else len(
                    objs) + summary[collection_name]

        print
        print "=" * 40
        print ' ' * 15, u'Stash'
        print "=" * 40
        print
        print u"数据已成功保存到MongoDB的theforce库中，其中新增数据:"
        for name, length in summary.items():
            print name, length
    finally:
        mongo.close()


def save(metas, batch_num=100):
    """
    读取配置，把Mongo数据同步到mysql中。
    """
    mongo = MongoClient(**config.mongo)
    db = MySQLdb.connect(**config.mysql)
    cursor = db.cursor()
    print
    print "=" * 40
    print ' ' * 15, u'Mongo --> MySQL'
    print "=" * 40
    print
    try:
        mongo_db = mongo.get_database('theforce')
        for meta in metas:
            for model_name, item_model in meta.iter_model():
                collection_name = item_model['name']
                table_name = item_model['table']
                attrs = meta.get_model_persist_attr_names(item_model)

                collection = mongo_db.get_collection(collection_name)
                results = [obj for obj in collection.find({})]
                sql = "insert into {0}({1}) values({2})".format(
                    table_name, ','.join(attrs),
                    ','.join(itertools.repeat('%s', len(attrs))))

                print
                print '-' * 40
                print u'开始处理{0}@mongo --> {1}@mysql, 共{2}条数据，每批{3}条批量迁移:'.format(
                    collection_name, table_name, len(results), batch_num)
                # 分组进行批量处理
                results2 = itertools.izip(itertools.count(), results)
                for group_key, group_it in itertools.groupby(
                        results2, lambda item: item[0] / batch_num):
                    print '.',
                    values = [[obj[attr] for attr in attrs]
                              for index, obj in group_it]
                    cursor.executemany(sql, values)
                print u'[完成]'
    finally:
        mongo.close()
        cursor.close()
        db.close()


if __name__ == "__main__":
    fn = sys.argv[1]
    paths = sys.argv[2:]
    metas = scan_data_dir(paths)

    if fn == 'check':
        load(metas)
    elif fn == 'stash':
        results = load(metas)
        stash(results)
    elif fn == 'save':
        save(metas)
