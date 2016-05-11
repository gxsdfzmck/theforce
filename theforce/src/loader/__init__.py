# -*- coding: utf-8 -*-
import os
import itertools

from openpyxl import load_workbook

from loader.helper import get_sheet_regions, get_sheets, get_evals
from logger import debug, error
from model import get_correct_rules_by_attr_label, get_attrs_with_validate_rules
from process import exec_correction, exec_evaluation, exec_validation


class ExcelLoader(object):
    """
    完成从Excel中载入各region数据，按(model_name, results)的顺序返回。

    1. 完成各值的自动纠错
    2. 添加ext值
    3. 计算evals值并自动纠错
    """

    def __init__(self, excel_abs_path, excel_config):
        self._abs_path = excel_abs_path
        self._path = os.path.basename(self._abs_path)
        self._workbook = load_workbook(filename=self._abs_path,
                                       data_only=True,
                                       read_only=True)
        self._config = excel_config

    def load(self):
        debug(u'开始解析文件:{0}'.format(self._path))
        for sheet_name, sheet_config in get_sheets(self._config):
            debug(u'开始解析Sheet:{0}'.format(sheet_name))
            try:
                sheet = self._workbook.get_sheet_by_name(sheet_name)
                yield [(name, region_results)
                       for name, region_results in self.load_sheet(
                           sheet, sheet_name, sheet_config)]
            except Exception, e:
                print e
                msg = u'在文件 %s 未找到Sheet: %s' % (self._path, sheet_name)
                error(msg)
                raise Exception(msg)

    def load_raw_data(self, sheet, sheet_name, sheet_config):
        # 加载原始数据，按区域返回
        for model_name, region, extend, fields, evals in get_sheet_regions(
                sheet_config):
            debug(u'读取[{0}]区域:{1}'.format(sheet_name, region))

            meta = {'model_name': model_name,
                      'region': region,
                      'extend': extend,
                      'fields': fields,
                      'evals': 'evals'}
            items = []
            for row in sheet.iter_rows(region):
                data = {'raw':[], 'eval':[], 'obj': {}}
                for i, col in enumerate(row):
                    data.raw.append({
                        'raw': col.value,
                        'row': col.row,
                        'col': col.col,
                        'key': fields[i]
                    })
                items.append(data)
            yield meta, items

    def verify_raw_data(self, meta, items):
        # 为原始数据生成校验内容
        # validate所有值
        self.validate(item, model_name)


    def load_sheet(self, sheet, sheet_name, sheet_config):
        for model_name, region, extend, fields, evals in get_sheet_regions(
                sheet_config):
            debug(u'读取[{0}]区域:{1}'.format(sheet_name, region))

            items = []
            for row in sheet.iter_rows(region):
                item = extend.copy() if extend else {}

                values = []
                row_no = 0
                for col in row:
                    values.append(col.value)
                    row_no = col.row

                item.update(dict([(field, value)
                                  for field, value in itertools.izip(fields,
                                                                     values)]))

                # 先correct收集到的值
                self.correct(item, model_name, fields)
                # 计算evals值并correct
                eval_fields = [attr_label
                               for attr_label, formulas in get_evals(evals)]
                self.evaluate(item, model_name, evals)
                self.correct(item, model_name, eval_fields)

                # validate所有值
                self.validate(item, model_name)
                # 载入debug信息
                self.pop_debug_info(item, sheet_name, row_no)
                items.append(item)
            yield model_name, items

    def pop_debug_info(self, item, sheet_name, row_no):
        item['__src'] = self._path
        item['__abs_src'] = self._abs_path
        item['__src_sheet'] = sheet_name
        item['__src_row'] = row_no
        return item

    def pop_info(self, attr, item, infos):
        if attr not in item:
            item[attr] = []
        if infos and len(infos) > 0:
            item[attr].extend(infos)

    def correct(self, item, model_name, attr_labels):
        for attr_label in attr_labels:
            init_value = item[attr_label]
            correction = get_correct_rules_by_attr_label(model_name,
                                                         attr_label)
            errors, warnings, value = exec_correction(attr_label, init_value,
                                                      correction)
            if value != init_value:
                item[attr_label] = value
            self.pop_info('__errors', item, errors)
            self.pop_info('__warnings', item, warnings)

    def evaluate(self, item, model_name, evalation):
        for attr_label, formulas in get_evals(evalation):
            errors, warnings, value = exec_evaluation(item, attr_label,
                                                      formulas)

            item[attr_label] = value
            self.pop_info('__errors', item, errors)
            self.pop_info('__warnings', item, warnings)

    def validate(self, item, model_name):
        for attr_label, attr_name, validation in get_attrs_with_validate_rules(
                model_name):
            errors = exec_validation(attr_label, item[attr_label] if
                                     attr_label in item else None, validation)
            self.pop_info('__errors', item, errors)
