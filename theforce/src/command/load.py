# -*- coding: utf-8 -*-
import os
import itertools

from openpyxl import load_workbook

from loader.helper import get_sheet_regions, get_sheets, get_evals
from logger import debug, error
from model import get_correct_rules_by_attr_label, get_attrs_with_validate_rules
from process import exec_correction, exec_evaluation, exec_validation


class ExcelLoader(object):
    """
    完成从Excel中载入各region数据，按(model_name, results)的顺序返回。
    """
    def __init__(self, excel_abs_path, excel_config):
        self._abs_path = excel_abs_path
        self._path = os.path.basename(self._abs_path)
        self._workbook = load_workbook(filename=self._abs_path,
                                       data_only=True,
                                       read_only=True)
        self._config = excel_config

    def load(self):
        debug(u'开始解析文件:{0}'.format(self._path))
        for sheet_name, sheet_config in get_sheets(self._config):
            debug(u'开始解析Sheet:{0}'.format(sheet_name))
            try:
                sheet = self._workbook.get_sheet_by_name(sheet_name)
                for meta, unit_datas in self.load_raw_data_by_sheet(
                        sheet, sheet_name, sheet_config):
                    yield meta, unit_datas
            except Exception, e:
                print e
                msg = u'在文件 %s 未找到Sheet: %s' % (self._path, sheet_name)
                error(msg)
                raise Exception(msg)

    def load_raw_data_by_sheet(self, sheet, sheet_name, sheet_config):
        # 加载原始数据，按区域返回
        for model_name, region, extend, fields, evals, validation, validation_row, correction_row, disable_validation, fields_correct_formula, sample, count in get_sheet_regions(
                sheet_config):
            debug(u'读取[{0}]区域:{1}'.format(sheet_name, region))

            meta = {'model_name': model_name,
                    'region': region,
                    'extend': extend,
                    'fields': fields,
                    'evals': evals,
                    'validation': validation,
                    'validation_row': validation_row,
                    'correction_row': correction_row,
                    'disable_validation': disable_validation,
                    'fields_correct_formula': fields_correct_formula,
                    'sample': sample,
                    'count': count,
                    '__src': self._path,
                    '__abs_src': self._abs_path,
                    '__src_sheet': sheet_name,
                    '__src_region': region}
            items = []
            for row in sheet.iter_rows(region):
                data = {'raw': [], 'eval': [], 'extend': [], 'obj': {}, 'missing': []}
                for i, col in enumerate(row):
                    data['raw'].append({
                        'row': col.row,
                        'col': col.column,
                        'value': col.value,
                        'key': fields[i]
                    })
                data['row'] = data['raw'][0]['row']
                items.append(data)
            yield meta, items
