# -*- coding: utf-8 -*-
import itertools
import os
from openpyxl import load_workbook
import shutil
import datetime
from logger import debug, error
from util import stringify


def save(raws, backup=True):
    # 先按文件分组
    for file_path, raw_group in itertools.groupby(
            raws, lambda item: item[0]['__abs_src']):
        modifies = [item for item in loop_changes(raw_group)]
        if len(modifies) > 0:
            if backup:
                file_name = os.path.basename(file_path)
                file_name = u"{0}_{1}_backup.xlsx".format(
                    os.path.splitext(file_name)[0],
                    datetime.datetime.now().strftime('%Y%m%d%H%M%S'))
                bk_file_path = os.path.join(
                    os.path.dirname(file_path), file_name)
                shutil.copyfile(file_path, bk_file_path)

            workbook = load_workbook(filename=file_path, data_only=True)

            debug(u'更新文件:{0}'.format(file_path))
            for meta, row_data, col_data in modifies:
                # 按区域更新
                name = meta['__src_sheet']
                row = col_data['row']
                col = col_data['col']
                raw_value = stringify(col_data['raw_value'])
                value = stringify(col_data['value'])
                sheet = workbook.get_sheet_by_name(name)
                sheet.cell(row=row, column=col).value = value
                debug(u'更新{0}-[行{1}:列{2}]：{3} --> {4}'.format(
                    name, row, col, raw_value, value))

            workbook.save(filename=file_path)


def loop_changes(raw_group):
    for meta, unit_datas in raw_group:
        for row_data in unit_datas:
            for col_data in row_data['raw']:
                if 'change' in col_data and col_data['change']:
                    if stringify(col_data['value']) != stringify(col_data[
                            'raw_value']):
                        yield meta, row_data, col_data
